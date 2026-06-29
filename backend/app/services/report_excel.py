import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel_report(
    industry: str,
    metrics: dict,
    eligibility: dict,
    roadmap: list,
) -> bytes:
    """Generate an Excel report with calculation results, eligibility, and roadmap."""

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fills = {
        "green": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
        "blue": PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
        "orange": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
        "purple": PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid"),
    }

    def style_header_row(ws, row, fill_key):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fills[fill_key]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")

    def style_data_row(ws, row):
        for cell in ws[row]:
            cell.border = thin_border

    ws = wb.active
    ws.title = "Carbon Report"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    ws.append(["AI Carbon Intelligence Platform - Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Industry", industry])
    ws["A3"].font = Font(bold=True)
    ws.append([])

    ws.append(["Input Metrics", "Value"])
    style_header_row(ws, ws.max_row, "green")
    rows = [
        ("Electricity (kWh)", f"{metrics.get('electricity_kwh', 0):,.2f}"),
        ("Diesel (Liters)", f"{metrics.get('fuel_diesel_liters', 0):,.2f}"),
        ("Waste (kg)", f"{metrics.get('waste_kg', 0):,.2f}"),
        ("Operational Hours", f"{metrics.get('operational_hours', 0):,.0f}"),
    ]
    for label, value in rows:
        ws.append([label, value])
        style_data_row(ws, ws.max_row)
    ws.append([])

    ws.append(["Emissions Summary", "Value (tCO2e)"])
    style_header_row(ws, ws.max_row, "blue")
    em_rows = [
        ("Scope 1 Emissions", f"{eligibility.get('scope_1', 0):.2f}"),
        ("Scope 2 Emissions", f"{eligibility.get('scope_2', 0):.2f}"),
        ("Total Emissions", f"{eligibility.get('total', 0):.2f}"),
    ]
    for label, value in em_rows:
        ws.append([label, value])
        style_data_row(ws, ws.max_row)
    ws.append([])

    ws.append(["Eligibility Score", "Value"])
    style_header_row(ws, ws.max_row, "orange")
    score_rows = [
        ("Readiness Score", f"{eligibility.get('readiness_score', 0)} / 100"),
        ("Emissions Rating", str(eligibility.get("emissions_rating", "N/A"))),
        ("Reduction Potential", f"{eligibility.get('reduction_potential_pct', 0):.1f}%"),
        ("Carbon Credit Potential", f"{eligibility.get('carbon_credit_potential', 0):.0f} credits"),
        ("Projected Revenue", f"INR {eligibility.get('projected_revenue_inr', 0):,.2f}"),
        ("Confidence Score", f"{eligibility.get('confidence_score', 0):.2f}"),
    ]
    for label, value in score_rows:
        ws.append([label, value])
        style_data_row(ws, ws.max_row)
    ws.append([])

    ws.append(["Year", "Recommendation", "Investment (INR)", "Savings (INR)", "Credits"])
    style_header_row(ws, ws.max_row, "purple")
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    for item in roadmap:
        ws.append([
            item.get("year", ""),
            item.get("recommendation", ""),
            f"{item.get('investment_inr', 0):,.0f}",
            f"{item.get('savings_inr', 0):,.0f}",
            f"{item.get('credits_earned', 0):.0f}",
        ])
        style_data_row(ws, ws.max_row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
